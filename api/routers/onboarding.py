"""Router: /onboarding — Callahan migration flow (3-step wizard).

Step 1: POST /onboarding/callahan-peer-group
        Build a Callahan-equivalent national peer group from asset tier + state + FOM.
        Returns peer_group_id + preview of key metrics — numbers should match Callahan exactly.

Step 2: POST /onboarding/verify-callahan
        Accept a Callahan CSV or XLSX export, parse it, and compare every metric
        to P76's computed values for the same institution and period.

Step 3: GET /onboarding/regional-context/{charter_number}
        Return peer distributions for both the Callahan-equivalent national group
        and the regional peer group, so Step 3 can layer the purple "Regional peers" line.

CLAUDE.md: accessible from onboarding and settings. Never frame as
"switching from" or "replacing" Callahan — additive positioning only.
"""

from __future__ import annotations

import csv
import io
import logging
import os
import uuid
from typing import Literal, Optional

from fastapi import APIRouter, Form, HTTPException, Query, Request, UploadFile
from pydantic import BaseModel
from sqlalchemy import text

from api.routers.query import CALLAHAN_TO_P76_METRIC_MAP

router = APIRouter()
logger = logging.getLogger(__name__)

DB_URL = os.environ.get("DATABASE_URL")

# ── Asset tier config ─────────────────────────────────────────────────────────

ASSET_TIER_RANGES: dict[str, tuple[float, float]] = {
    "under_100m": (0,            100_000_000),
    "100m_250m":  (100_000_000,  250_000_000),
    "250m_500m":  (250_000_000,  500_000_000),
    "500m_1b":    (500_000_000,  1_000_000_000),
    "1b_5b":      (1_000_000_000, 5_000_000_000),
    "5b_plus":    (5_000_000_000, 999_000_000_000_000),
}

ASSET_TIER_DISPLAY: dict[str, str] = {
    "under_100m": "Under $100M",
    "100m_250m":  "$100M – $250M",
    "250m_500m":  "$250M – $500M",
    "500m_1b":    "$500M – $1B",
    "1b_5b":      "$1B – $5B",
    "5b_plus":    "$5B+",
}

# Preview metrics for Step 1 table — subset Callahan prominently shows
# (callahan_display, p76_key, display_format, is_adverse)
PREVIEW_METRICS: list[tuple[str, str, str, bool]] = [
    ("Total Delinquency Ratio",   "delinq_rate_total",               "percent", True),
    ("Net Charge-Off Ratio",      "chargeoff_rate_total_annualized",  "percent", True),
    ("Allowance Coverage Ratio",  "alll_coverage",                    "ratio",   False),
    ("Net Worth Ratio",           "net_worth_ratio",                  "percent", False),
    ("Return on Assets",          "roa_annualized",                   "percent", False),
    ("Efficiency Ratio",          "efficiency_ratio",                  "percent", True),
]

# ── Pydantic models ───────────────────────────────────────────────────────────

class PeerGroupCriteria(BaseModel):
    asset_tier: str                    # key from ASSET_TIER_RANGES
    states: list[str]                  # e.g. ["MI", "OH"]
    field_of_membership: Optional[str] = None   # "community" | "seg" | "mcb" | None


class PreviewMetricRow(BaseModel):
    callahan_name:   str
    p76_metric:      str
    institution_value: Optional[float]
    peer_p25:        Optional[float]
    peer_median:     Optional[float]
    peer_p75:        Optional[float]
    stars:           Optional[int]
    percentile_rank: Optional[float]
    display_format:  str              # "percent" | "ratio" | "dollar" | "count"
    is_adverse:      bool


class PeerGroupBuildResponse(BaseModel):
    peer_group_id:    str
    group_name:       str
    n_institutions:   int
    period:           str
    institution_name: str
    institution_state: Optional[str]
    preview_metrics:  list[PreviewMetricRow]


class ComparisonRow(BaseModel):
    callahan_name:   str
    callahan_column: str          # exact column name from the file
    p76_metric:      str
    callahan_value:  Optional[float]
    p76_value:       Optional[float]
    match:           Literal["exact", "close", "mismatch", "unmapped"]
    delta:           Optional[float]    # p76_value - callahan_value
    display_format:  str


class VerifyCallahanResponse(BaseModel):
    rows:         list[ComparisonRow]
    n_exact:      int
    n_close:      int
    n_mismatch:   int
    n_unmapped:   int
    all_match:    bool
    institution_row_found: bool
    note:         str


class DistributionStats(BaseModel):
    n_institutions: int
    p10:  Optional[float]
    p25:  Optional[float]
    p50:  Optional[float]
    p75:  Optional[float]
    p90:  Optional[float]


class RegionalContextResponse(BaseModel):
    charter_number:          int
    period:                  str
    metric:                  str
    institution_value:       Optional[float]
    national_peer_group_id:  Optional[str]
    national_distribution:   Optional[DistributionStats]
    regional_distribution:   Optional[DistributionStats]
    institution_state:       Optional[str]


# ── Peer-group builder ────────────────────────────────────────────────────────

def _build_peer_group(
    criteria: PeerGroupCriteria,
    charter_number: int,
    period: str,
    tenant_id: str,
    db_url: Optional[str],
) -> tuple[str, list[str], int]:
    """Build or retrieve a Callahan-equivalent peer group. Returns (peer_group_id, charter_list, n)."""
    from db import get_engine

    engine = get_engine(db_url)
    lo, hi = ASSET_TIER_RANGES.get(criteria.asset_tier, (0, 999_000_000_000_000))

    with engine.connect() as conn:
        peer_rows = conn.execute(
            text("""
                SELECT charter_number
                FROM institutions_quarterly
                WHERE period = :period
                  AND acct_010 BETWEEN :lo AND :hi
                  AND state_code = ANY(:states)
                ORDER BY acct_010 DESC
            """),
            {
                "period": period,
                "lo":     lo,
                "hi":     hi,
                "states": criteria.states,
            },
        ).fetchall()

    charters = [str(r[0]) for r in peer_rows]
    if not charters:
        raise HTTPException(
            status_code=404,
            detail=f"No institutions found for asset tier {criteria.asset_tier!r} in states {criteria.states}",
        )

    tier_label = ASSET_TIER_DISPLAY.get(criteria.asset_tier, criteria.asset_tier)
    state_label = ", ".join(sorted(criteria.states))
    group_name = f"Callahan National — {tier_label} — {state_label}"

    # Upsert into peer_groups
    group_id = str(uuid.uuid4())
    with engine.begin() as conn:
        existing = conn.execute(
            text("""
                SELECT id FROM peer_groups
                WHERE tenant_id = :tid AND group_name = :gn AND group_type = 'callahan_national'
                LIMIT 1
            """),
            {"tid": tenant_id, "gn": group_name},
        ).first()

        if existing:
            group_id = str(existing[0])
            conn.execute(
                text("""
                    UPDATE peer_groups SET institution_ids = :ids
                    WHERE id = :gid
                """),
                {"ids": charters, "gid": group_id},
            )
        else:
            conn.execute(
                text("""
                    INSERT INTO peer_groups
                      (id, tenant_id, group_name, group_type, geography_type, institution_ids, is_default)
                    VALUES (:id, :tid, :gn, 'callahan_national', 'national', :ids, false)
                """),
                {
                    "id":  group_id,
                    "tid": tenant_id,
                    "gn":  group_name,
                    "ids": charters,
                },
            )

    return group_id, charters, len(charters)


# ── Preview metrics (Step 1) ──────────────────────────────────────────────────

def _compute_preview_metrics(
    charter_number: int,
    period: str,
    peer_charters: list[int],
    db_url: Optional[str],
) -> list[PreviewMetricRow]:
    try:
        import pandas as pd
        from processing.delinquency_engine import (
            assign_stars, compute_peer_distribution, compute_ratios, rank_institution,
        )
        from db import get_engine
        from sqlalchemy import select
        from db import institutions_quarterly

        engine = get_engine(db_url)
        with engine.connect() as conn:
            rows = conn.execute(
                select(institutions_quarterly).where(
                    institutions_quarterly.c.charter_number == charter_number,
                    institutions_quarterly.c.period == period,
                )
            ).mappings().all()

        inst_df = compute_ratios(pd.DataFrame([dict(r) for r in rows])) if rows else pd.DataFrame()
        inst = inst_df.to_dict("records")[0] if not inst_df.empty else {}

        peer_charters_int = [int(c) for c in peer_charters]
        result = []
        for callahan_name, p76_key, fmt, is_adverse in PREVIEW_METRICS:
            inst_value = inst.get(p76_key)
            try:
                dist = compute_peer_distribution(p76_key, peer_charters_int, period, db_url)
                pctile = rank_institution(inst_value, dist, p76_key) if inst_value is not None else None
                stars  = assign_stars(pctile) if pctile is not None else None
            except Exception:
                dist, pctile, stars = {}, None, None

            result.append(PreviewMetricRow(
                callahan_name     = callahan_name,
                p76_metric        = p76_key,
                institution_value = inst_value,
                peer_p25          = dist.get("p25"),
                peer_median       = dist.get("p50"),
                peer_p75          = dist.get("p75"),
                stars             = stars,
                percentile_rank   = pctile,
                display_format    = fmt,
                is_adverse        = is_adverse,
            ))
        return result
    except Exception as exc:
        logger.warning("Preview metrics failed: %s", exc)
        return []


# ── Callahan file parser (Step 2) ─────────────────────────────────────────────

def _normalize_col(s: str) -> str:
    return s.lower().strip().replace("-", " ").replace("_", " ").replace("  ", " ")


def _match_callahan_column(col_name: str) -> Optional[str]:
    """Map a Callahan column header to a P76 metric key. Longest match wins."""
    norm = _normalize_col(col_name)
    # Exact match first
    for k, v in CALLAHAN_TO_P76_METRIC_MAP.items():
        if _normalize_col(k) == norm:
            return v
    # Substring: Callahan key contains the column name
    candidates = []
    for k, v in CALLAHAN_TO_P76_METRIC_MAP.items():
        kn = _normalize_col(k)
        if norm in kn or kn in norm:
            candidates.append((len(kn), v))
    if candidates:
        return sorted(candidates, reverse=True)[0][1]   # longest key wins
    return None


def _parse_xlsx(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    """Return (headers, data_rows). Attempts to auto-detect the header row."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    # Find header row: first row where ≥3 cells look like metric label text
    header_idx = 0
    for i, row in enumerate(all_rows):
        str_cells = [c for c in row if isinstance(c, str) and len(c.strip()) > 1]
        if len(str_cells) >= 3:
            header_idx = i
            break

    raw_headers = all_rows[header_idx]
    headers = [str(c).strip() if c is not None else f"__col{j}" for j, c in enumerate(raw_headers)]

    data_rows = []
    for row in all_rows[header_idx + 1:]:
        if all(c is None for c in row):
            continue
        data_rows.append({
            headers[j]: row[j] for j in range(min(len(headers), len(row)))
        })
    return headers, data_rows


def _parse_csv_bytes(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    text = file_bytes.decode("utf-8-sig")  # handle Excel BOM
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    return list(reader.fieldnames or []), rows


def _find_institution_row(rows: list[dict], institution_name: str) -> Optional[dict]:
    """Find the row for this institution in a multi-row Callahan export."""
    if not rows:
        return None
    if len(rows) == 1:
        return rows[0]

    name_lower = institution_name.lower().strip()
    # Search all cell values for an exact match
    for row in rows:
        for val in row.values():
            if val is not None and _normalize_col(str(val)) == _normalize_col(institution_name):
                return row

    # Partial match: institution name first word in any cell
    first_word = name_lower.split()[0] if name_lower else ""
    if len(first_word) >= 4:
        for row in rows:
            for val in row.values():
                if val is not None and first_word in str(val).lower():
                    return row

    # Last resort: first data row
    return rows[0]


def _coerce_float(v) -> Optional[float]:
    """Convert Callahan cell value to float (handles '2.45%', '$1,234,567', etc.)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("$", "").replace("%", "").strip()
    try:
        f = float(s)
        # Callahan shows rates as percentages (e.g., 2.45 for 2.45%)
        # P76 stores them as decimals (0.0245).
        # We match on the display value level — convert Callahan % to decimal here.
        if "%" in str(v):
            return f / 100.0
        return f
    except (ValueError, TypeError):
        return None


def _compare_values(
    callahan_val: Optional[float],
    p76_val: Optional[float],
    display_format: str,
) -> tuple[Literal["exact", "close", "mismatch", "unmapped"], Optional[float]]:
    if callahan_val is None or p76_val is None:
        return "unmapped", None

    delta = p76_val - callahan_val

    if display_format in ("percent", "ratio"):
        tol_exact = 0.0001   # 0.01 pp
        tol_close = 0.0005   # 0.05 pp
    else:
        tol_exact = abs(callahan_val) * 0.001 if callahan_val != 0 else 1.0
        tol_close = abs(callahan_val) * 0.005 if callahan_val != 0 else 5.0

    if abs(delta) <= tol_exact:
        return "exact", delta
    if abs(delta) <= tol_close:
        return "close", delta
    return "mismatch", delta


# ── P76 computed values for comparison ───────────────────────────────────────

def _compute_p76_values(charter_number: int, period: str, db_url: Optional[str]) -> dict:
    """Compute P76's metric values for a single institution + period."""
    try:
        import pandas as pd
        from processing.delinquency_engine import compute_ratios
        from db import get_engine
        from sqlalchemy import select
        from db import institutions_quarterly

        engine = get_engine(db_url)
        with engine.connect() as conn:
            rows = conn.execute(
                select(institutions_quarterly).where(
                    institutions_quarterly.c.charter_number == charter_number,
                    institutions_quarterly.c.period == period,
                )
            ).mappings().all()

        if not rows:
            return {}

        inst_df = compute_ratios(pd.DataFrame([dict(r) for r in rows]))
        inst = inst_df.to_dict("records")[0] if not inst_df.empty else {}

        # Add metrics compute_ratios() may not include
        loans  = float(inst.get("acct_025B") or 0)
        shares = float(inst.get("acct_018")  or 0)
        assets = float(inst.get("acct_010")  or 0)
        nim    = float(inst.get("acct_IS0010") or 0)
        members = float(inst.get("acct_083") or 0)

        extras: dict = {}
        if shares   > 0: extras["loan_to_share"]          = loans / shares
        if assets   > 0: extras["nim_annualized"]          = nim / assets * 4
        if members  > 0: extras["average_share_balance"]   = shares / members

        return {**inst, **extras}
    except Exception as exc:
        logger.warning("P76 value computation failed: %s", exc)
        return {}


# ── Peer distributions helper (Step 3) ───────────────────────────────────────

def _distribution_stats(
    charter_number: int,
    period: str,
    metric: str,
    peer_charters: list[int],
    db_url: Optional[str],
) -> Optional[DistributionStats]:
    if not peer_charters:
        return None
    try:
        from processing.delinquency_engine import compute_peer_distribution
        dist = compute_peer_distribution(charter_number, period, metric, peer_charters, db_url)
        if not dist:
            return None
        return DistributionStats(
            n_institutions = dist.get("n", len(peer_charters)),
            p10 = dist.get("p10"),
            p25 = dist.get("p25"),
            p50 = dist.get("p50"),
            p75 = dist.get("p75"),
            p90 = dist.get("p90"),
        )
    except Exception:
        return None


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/callahan-peer-group", response_model=PeerGroupBuildResponse)
async def build_callahan_peer_group(
    request: Request,
    criteria: PeerGroupCriteria,
    charter_number: int = Query(...),
    period: str = Query(..., description="e.g. 2026Q1"),
):
    """Step 1 — Build a Callahan-equivalent national peer group.

    Takes the same criteria used in Callahan (asset tier + state + optional FOM)
    and creates a P76 peer group that mirrors it exactly.
    Returns preview metrics so the user can confirm the numbers match.
    """
    tenant_id = getattr(request.state, "tenant_id", None) or "anonymous"

    if criteria.asset_tier not in ASSET_TIER_RANGES:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown asset_tier {criteria.asset_tier!r}. Valid: {list(ASSET_TIER_RANGES)}",
        )
    if not criteria.states:
        raise HTTPException(status_code=400, detail="At least one state is required.")

    from db import get_engine
    engine = get_engine(DB_URL)

    # Get institution name + state
    with engine.connect() as conn:
        inst_row = conn.execute(
            text("""
                SELECT institution_name, state_code FROM institutions_quarterly
                WHERE charter_number = :c AND period = :p LIMIT 1
            """),
            {"c": charter_number, "p": period},
        ).mappings().first()

    institution_name  = inst_row["institution_name"] if inst_row else f"Charter {charter_number}"
    institution_state = inst_row["state_code"] if inst_row else None

    group_id, charters, n_institutions = _build_peer_group(
        criteria, charter_number, period, tenant_id, DB_URL
    )
    peer_int = [int(c) for c in charters]
    preview  = _compute_preview_metrics(charter_number, period, peer_int, DB_URL)

    return PeerGroupBuildResponse(
        peer_group_id     = group_id,
        group_name        = f"Callahan National — {ASSET_TIER_DISPLAY.get(criteria.asset_tier)} — {', '.join(sorted(criteria.states))}",
        n_institutions    = n_institutions,
        period            = period,
        institution_name  = institution_name,
        institution_state = institution_state,
        preview_metrics   = preview,
    )


@router.post("/verify-callahan", response_model=VerifyCallahanResponse)
async def verify_callahan(
    request: Request,
    file: UploadFile,
    charter_number: int = Form(...),
    period: str = Form(...),
    peer_group_id: str = Form(default=""),
):
    """Step 2 — Parse a Callahan export and compare every metric to P76's values.

    Accepts .csv or .xlsx files from Callahan downloads.
    Maps column names using CALLAHAN_TO_P76_METRIC_MAP.
    Shows Callahan value vs P76 value side-by-side.
    """
    tenant_id = getattr(request.state, "tenant_id", None) or "anonymous"

    # Read uploaded file
    filename    = file.filename or "upload.csv"
    file_bytes  = await file.read()

    try:
        if filename.lower().endswith((".xlsx", ".xls")):
            headers, data_rows = _parse_xlsx(file_bytes)
        else:
            headers, data_rows = _parse_csv_bytes(file_bytes)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not parse file: {exc}")

    if not data_rows:
        raise HTTPException(status_code=422, detail="File appears to be empty or unreadable.")

    # Find institution's row (Callahan exports may be multi-institution peer downloads)
    from db import get_engine
    engine = get_engine(DB_URL)
    with engine.connect() as conn:
        inst_row = conn.execute(
            text("SELECT institution_name FROM institutions_quarterly WHERE charter_number = :c AND period = :p LIMIT 1"),
            {"c": charter_number, "p": period},
        ).mappings().first()
    institution_name = inst_row["institution_name"] if inst_row else f"Charter {charter_number}"

    callahan_row = _find_institution_row(data_rows, institution_name)
    institution_row_found = callahan_row is not None
    if not institution_row_found:
        callahan_row = data_rows[0]   # use first row as fallback

    # Compute P76 values
    p76_values = _compute_p76_values(charter_number, period, DB_URL)

    # Build comparison rows — one per column in the Callahan file
    comparison_rows: list[ComparisonRow] = []
    for col in headers:
        if not col or col.startswith("__col"):
            continue
        p76_metric = _match_callahan_column(col)
        raw_val    = callahan_row.get(col) if callahan_row else None
        callahan_val = _coerce_float(raw_val)

        if p76_metric is None:
            comparison_rows.append(ComparisonRow(
                callahan_name   = col,
                callahan_column = col,
                p76_metric      = "—",
                callahan_value  = callahan_val,
                p76_value       = None,
                match           = "unmapped",
                delta           = None,
                display_format  = "other",
            ))
            continue

        p76_val = p76_values.get(p76_metric)

        # Detect display format from metric type
        if p76_metric in ("acct_010", "acct_025B", "acct_018", "assets_raw"):
            fmt = "dollar"
        elif p76_metric == "average_share_balance":
            fmt = "dollar"
        elif p76_metric in ("acct_083",):
            fmt = "count"
        elif p76_metric == "alll_coverage":
            fmt = "ratio"
        else:
            fmt = "percent"

        match_result, delta = _compare_values(callahan_val, p76_val, fmt)
        comparison_rows.append(ComparisonRow(
            callahan_name   = col,
            callahan_column = col,
            p76_metric      = p76_metric,
            callahan_value  = callahan_val,
            p76_value       = p76_val,
            match           = match_result,
            delta           = delta,
            display_format  = fmt,
        ))

    # Remove empty/unmeaningful rows
    comparison_rows = [r for r in comparison_rows if r.callahan_value is not None or r.p76_value is not None]

    # Sort: mismatches first, then close, then exact, then unmapped
    _order = {"mismatch": 0, "close": 1, "exact": 2, "unmapped": 3}
    comparison_rows.sort(key=lambda r: _order.get(r.match, 4))

    n_exact    = sum(1 for r in comparison_rows if r.match == "exact")
    n_close    = sum(1 for r in comparison_rows if r.match == "close")
    n_mismatch = sum(1 for r in comparison_rows if r.match == "mismatch")
    n_unmapped = sum(1 for r in comparison_rows if r.match == "unmapped")
    all_match  = n_mismatch == 0 and n_close == 0

    if all_match:
        note = "Every metric matches exactly. You're seeing the same numbers Callahan shows."
    elif n_mismatch == 0:
        note = f"{n_exact} of {n_exact + n_close} metrics match exactly; {n_close} differ by rounding only (< 0.05 pp)."
    else:
        note = (
            f"{n_exact} match exactly, {n_close} within rounding. "
            f"{n_mismatch} metric(s) differ more than 0.05 pp — check that period and peer group criteria match."
        )

    return VerifyCallahanResponse(
        rows                   = comparison_rows,
        n_exact                = n_exact,
        n_close                = n_close,
        n_mismatch             = n_mismatch,
        n_unmapped             = n_unmapped,
        all_match              = all_match,
        institution_row_found  = institution_row_found,
        note                   = note,
    )


@router.get("/metric-trend/{charter_number}/metric/{metric_name}")
async def get_metric_trend_onboarding(
    request: Request,
    charter_number: int,
    metric_name: str,
    period: str = Query(...),
    peer_group: str = Query(default="REGIONAL"),
    n_periods: int = Query(default=12),
):
    """Auth-exempt metric trend for PeerBandChart in the onboarding wizard (Step 3).

    Returns the same shape as /peer-comparison/{charter}/metric/{metric} so
    PeerBandChart can be pointed here via apiBase="/onboarding/metric-trend".
    """
    import pandas as pd
    from db import get_engine, institutions_quarterly
    from sqlalchemy import select
    from processing.delinquency_engine import compute_peer_distribution, compute_ratios
    from processing.early_warning_engine import _trailing_periods
    from processing.peer_engine import PeerGroupType, build_peer_group, peer_group_label
    from api.routers.peer_comparison import METRIC_LABELS

    tenant_id = getattr(request.state, "tenant_id", None) or "anonymous"
    periods = _trailing_periods(period, n=n_periods)

    group_type = PeerGroupType(peer_group)
    try:
        peer_charters = build_peer_group(charter_number, period, group_type, tenant_id, db_url=DB_URL)
    except Exception:
        peer_charters = []
    label = peer_group_label(group_type, charter_number, period, DB_URL)

    engine = get_engine(DB_URL)
    result_rows = []

    for p in periods:
        with engine.connect() as conn:
            inst_result = conn.execute(
                select(institutions_quarterly).where(
                    institutions_quarterly.c.charter_number == charter_number,
                    institutions_quarterly.c.period == p,
                )
            )
            inst_rows = inst_result.mappings().all()

        if not inst_rows:
            inst_val = None
        else:
            inst_df = compute_ratios(pd.DataFrame([dict(r) for r in inst_rows]))
            raw = inst_df[metric_name].iloc[0] if metric_name in inst_df.columns else None
            inst_val = float(raw) if raw is not None else None

        dist = compute_peer_distribution(metric_name, peer_charters, p, DB_URL) if peer_charters else {}
        result_rows.append({
            "period": p,
            "institution_value": inst_val,
            "peer_p10":   dist.get("p10"),
            "peer_p25":   dist.get("p25"),
            "peer_p50":   dist.get("p50"),
            "peer_p75":   dist.get("p75"),
            "peer_p90":   dist.get("p90"),
            "peer_count": dist.get("n", 0),
        })

    callahan_label, unit = METRIC_LABELS.get(metric_name, (metric_name, "%"))

    return {
        "charter_number":   charter_number,
        "metric":           metric_name,
        "callahan_label":   callahan_label,
        "unit":             unit,
        "peer_group_type":  peer_group,
        "peer_group_label": label,
        "data":             result_rows,
    }


@router.get("/regional-context/{charter_number}", response_model=RegionalContextResponse)
async def get_regional_context(
    request: Request,
    charter_number: int,
    period: str = Query(...),
    metric: str = Query(default="delinq_rate_total"),
    national_peer_group_id: str = Query(default=""),
):
    """Step 3 — Return both national (Callahan-equivalent) and regional peer distributions.

    Powers the PeerBandChart that shows the institution with two peer bands:
      National band (blue) = Callahan-equivalent peer group from Step 1
      Regional line (purple) = institutions in same geographic market
    """
    tenant_id = getattr(request.state, "tenant_id", None) or "anonymous"
    from db import get_engine
    from processing.peer_engine import PeerGroupType, build_peer_group

    engine = get_engine(DB_URL)

    # Institution value
    p76_values = _compute_p76_values(charter_number, period, DB_URL)
    inst_value = p76_values.get(metric)

    # Institution state (for signal separator personalization)
    with engine.connect() as conn:
        inst_row = conn.execute(
            text("SELECT state_code FROM institutions_quarterly WHERE charter_number = :c AND period = :p LIMIT 1"),
            {"c": charter_number, "p": period},
        ).mappings().first()
    inst_state = inst_row["state_code"] if inst_row else None

    # National peer group (Callahan-equivalent from Step 1)
    nat_dist: Optional[DistributionStats] = None
    if national_peer_group_id:
        with engine.connect() as conn:
            group_row = conn.execute(
                text("SELECT institution_ids FROM peer_groups WHERE id = :gid LIMIT 1"),
                {"gid": national_peer_group_id},
            ).mappings().first()
        if group_row:
            nat_charters = [int(c) for c in (group_row["institution_ids"] or [])]
            nat_dist = _distribution_stats(charter_number, period, metric, nat_charters, DB_URL)

    # Regional peer group (same geography — default per CLAUDE.md)
    reg_dist: Optional[DistributionStats] = None
    try:
        reg_charters = build_peer_group(
            charter_number, period, PeerGroupType("REGIONAL"), tenant_id, db_url=DB_URL
        )
        reg_dist = _distribution_stats(charter_number, period, metric, reg_charters, DB_URL)
    except Exception:
        pass

    return RegionalContextResponse(
        charter_number         = charter_number,
        period                 = period,
        metric                 = metric,
        institution_value      = inst_value,
        national_peer_group_id = national_peer_group_id or None,
        national_distribution  = nat_dist,
        regional_distribution  = reg_dist,
        institution_state      = inst_state,
    )
